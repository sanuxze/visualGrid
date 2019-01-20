from copy import deepcopy
from openpyxl.styles import Border, Side, PatternFill


class VisualGridMaker(object):

    def __init__(self, info, worksheet, win_lines):
        self.ws = worksheet
        self.info = info
        self.win_lines = win_lines
        self.xcel_info = {
            'cell_start_row': 3,
            'cell_start_column': 1,
            'cell_end_row': 2 + info['view_size'],
            'cell_end_column': info['num_reels'],
            'shift': 2 + info['view_size']
        }

    @staticmethod
    def set_border(worksheet, cell_range):
        rows = worksheet[cell_range]
        side_thick = Side(border_style='thick', color="FF000000")
        side_thin = Side(border_style='thin', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

                # applying the thin borders for the symbol grid
                border.bottom = border.top = border.right = border.left = side_thin
                cell.border = border

                # applying the thick borders around the grid
                if pos_x == 0:
                    border.left = side_thick
                if pos_x == max_x:
                    border.right = side_thick
                if pos_y == 0:
                    border.top = side_thick
                if pos_y == max_y:
                    border.bottom = side_thick

                # set new border only if it's one of the edge cells
                # if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border

    def grid_maker(self):
        # creating a sample list like below as per the symbol grid:
        # sample = [['\t', '\t', '\t', '\t', '\t'],
        #           ['\t', '\t', '\t', '\t', '\t'],
        #           ['\t', '\t', '\t', '\t', '\t']]
        sample = []
        for size in range(self.info['view_size']):
            _line = []
            for reel in range(self.info['num_reels']):
                _line.append("\t")
            sample.append(_line)

        # print(sample)

        # making the required grid from sample list
        grids_for_xcel = []
        for _line in self.win_lines:
            copy_sample = deepcopy(sample)
            for _col, _row in enumerate(_line):
                copy_sample[_row][_col] = 'X'
            grids_for_xcel.append(copy_sample)

        # print(grids_for_xcel)

        # writing the grid to worksheet
        for _index, _grid in enumerate(grids_for_xcel):
            self.ws.append(["\n"])
            self.ws.append(["Line_" + str(_index + 1)])
            for _line in _grid:
                self.ws.append(_line)

        # applying the border to the grid
        shift = 0
        for _grid in range(len(grids_for_xcel)):
            # cell start and end need to be given
            cell_start = self.ws.cell(row=self.xcel_info['cell_start_row'] + shift,
                                      column=self.xcel_info['cell_start_column'])
            cell_end = self.ws.cell(row=self.xcel_info['cell_end_row'] + shift,
                                    column=self.xcel_info['cell_end_column'])

            shift += self.xcel_info['shift']

            # print(cell_start.coordinate, cell_end.coordinate)
            self.set_border(self.ws, "{}:{}".format(cell_start.coordinate, cell_end.coordinate))

        # applying the fill for every X it finds
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value == "X":
                    # apply color fill
                    red_fill = PatternFill(start_color='FFFF00',
                                           end_color='FFFF00',
                                           fill_type='solid')

                    self.ws[cell.coordinate].fill = red_fill

        # dm = dimensions.ColumnDimension(worksheet=self.ws, index='A', bestFit=True)
